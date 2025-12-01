from customtkinter import *
from tkinter import messagebox, ttk, filedialog
from tkcalendar import DateEntry
from Form import BaseForm
import openpyxl
from openpyxl.styles import Font, Alignment

class Create_DatCho(CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color="#FFFFFF")
        
        self.db = BaseForm.ConnectionDatabase()
        self.Create_frameBottom()
        self.Create_frameTop()
        self.load_data()
        
        self.list_them = []
        self.list_xoa = []
        self.list_sua = []
        
    def Create_frameBottom(self):
        self.frameBottom = CTkFrame(self, height=400, fg_color="#FFFFFF")
        self.frameBottom.pack(side="bottom", fill="x")
        
        columns = ("MaDatCho", "MaKhachHang", "MaNhanVien", "MaTour", "SoLuongNguoiLon", 
                  "SoLuongTreEm", "TongTien", "NgayDat", "TrangThaiBooking")
        
        self.tree = ttk.Treeview(self.frameBottom, columns=columns, show="headings")
        
        scrollbar_y = CTkScrollbar(self.frameBottom, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar_y.set)
        
        scrollbar_x = CTkScrollbar(self.frameBottom, command=self.tree.xview, orientation="horizontal")
        self.tree.configure(xscrollcommand=scrollbar_x.set)
        
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)
        style = ttk.Style()
        style.theme_use("default")
        style.configure(
            "Treeview.Heading",
            background="#244f88",
            foreground="white",
            font=("Segoe UI", 10, "bold")
        )
        style.map(
            "Treeview.Heading",
            background=[("active", "#1a5bb8")]
        )
        
        cols = [
            ("MaDatCho", "Mã Đặt Chỗ", 150, "center"),
            ("MaKhachHang", "Mã khách hàng", 150, "center"),
            ("MaNhanVien", "Mã nhân viên", 150, "center"),
            ("MaTour", "Mã Tour", 150, "center"),
            ("SoLuongNguoiLon", "Người lớn", 90, "center"),
            ("SoLuongTreEm", "Trẻ em", 80, "center"),
            ("TongTien", "Tổng tiền", 150, "e"),
            ("NgayDat", "Ngày đặt", 150, "center"),
            ("TrangThaiBooking", "Trạng thái", 150, "center"),
        ]
        for col, text, width, anchor in cols:
            self.tree.heading(col, text=text)
            self.tree.column(col, width=width, anchor=anchor, stretch=False)
    
    def Create_frameTop(self):
        self.frameTop = CTkFrame(self,width=400, height=300, fg_color="#FFFFFF")
        self.frameTop.pack(side="top", fill="both", expand=True)
        
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
            
        self.lb_TieuDe = CTkLabel(self.frameTop, text="Thông tin chi tiết của Đặt chỗ", font=("Segoe UI", 17, "bold"))
        self.lb_TieuDe.place(x=20, y=65)
    #MaDatCho
        self.lb_MaDatCho = CTkLabel(self.frameTop, text= "Mã đặt chỗ", font=("Segoe UI", 14))
        self.lb_MaDatCho.place(x=20, y = 100) 
        self.entry_MaDatCho = CTkEntry(self.frameTop, width=250, height=20)
        self.entry_MaDatCho.place(x=120, y=105)
    #MaKhachHang
        self.lb_MaKhachHang = CTkLabel(self.frameTop, text= "Mã khách hàng", font=("Segoe UI", 14))
        self.lb_MaKhachHang.place(x=20, y = 140)
        self.cb_MaKhachHang = CTkComboBox(self.frameTop, width=250, height=20)
        self.cb_MaKhachHang.place(x=120, y=145)              
    #MaNhanVien
        self.lb_MaNhanVien = CTkLabel(self.frameTop, text= "Mã nhân viên", font=("Segoe UI", 14))
        self.lb_MaNhanVien.place(x=20, y = 180)
        self.cb_MaNhanVien = CTkComboBox(self.frameTop, width=250, height=20)
        self.cb_MaNhanVien.place(x=120, y=185)
    #MaTour
        self.lb_MaTour = CTkLabel(self.frameTop, text= "Mã Tour", font=("Segoe UI", 14))
        self.lb_MaTour.place(x=20, y = 220)
        self.cb_MaTour = CTkComboBox(self.frameTop, width=250, height=20)
        self.cb_MaTour.place(x=120, y=225)        
    #TrangThai
        self.lb_TrangThai = CTkLabel(self.frameTop, text= "Trạng thái", font=("Segoe UI", 14))
        self.lb_TrangThai.place(x=20, y = 260)
        self.cb_TrangThai = CTkComboBox(self.frameTop, width=250, height=20)
        self.cb_TrangThai.place(x=120, y=265)      
          
    #SoLuongTreEm
        self.lb_SoLuongTreEm = CTkLabel(self.frameTop, text= "Số lượng trẻ em", font=("Segoe UI", 14))
        self.lb_SoLuongTreEm.place(x=400, y = 100)
        self.entry_SoLuongTreEm = CTkEntry(self.frameTop, width=200, height=20)
        self.entry_SoLuongTreEm.place(x=555, y=105)
    #SoLuongNguoiLon
        self.lb_SoLuongNguoiLon = CTkLabel(self.frameTop, text= "Số lượng người lớn", font=("Segoe UI", 14))
        self.lb_SoLuongNguoiLon.place(x=400, y = 140)
        self.entry_SoLuongNguoiLon = CTkEntry(self.frameTop, width=200, height=20)
        self.entry_SoLuongNguoiLon.place(x=555, y=145)
    #NgayDat
        self.lb_NgayDat = CTkLabel(self.frameTop, text= "Ngày đặt", font=("Segoe UI", 14))
        self.lb_NgayDat.place(x=400, y = 180)
        self.date_NgayDat = DateEntry(self.frameTop, width=38, height=20)
        self.date_NgayDat.place(x=510, y=185) 
    #TongTien
        self.lb_TongTien = CTkLabel(self.frameTop, text= "Tổng tiền", font=("Segoe UI", 14))
        self.lb_TongTien.place(x=415, y = 220)
        self.entry_TongTien = CTkEntry(self.frameTop, width=250, height=20, fg_color="#928FA7")
        self.entry_TongTien.place(x=510, y=225)  
        self.entry_TongTien.configure(state="disabled")       

#--------------------------------------
# CHỨC NĂNG TÌM KIẾM
#--------------------------------------    
        self.cb_TimKiem = CTkComboBox(self.frameTop, width=130, height= 20)
        self.cb_TimKiem.place(x=380, y = 315)
        
        self.entry_TimKiem = CTkEntry(self.frameTop, width=330, height=20, fg_color="#FFFFFF")
        self.entry_TimKiem.place(x =520, y=315)
        
        self.btn_Timkiem = CTkButton(self.frameTop, width=40, height=14, text="🔍",
                                     font=("Segoe UI", 14, "bold"),
                                     text_color="#FFFFFF", command=self.TimKiem)
        self.btn_Timkiem.place(x=855,y=314)   
#--------------------------------------
# TẠO CÁC BUTTON
#--------------------------------------    

        if BaseForm.UserSession.is_admin():
    #Thêm
            self.btn_Them = CTkButton(self.frameTop, width=70, height=25, text="➕ Thêm",
                                    fg_color="#1D8D13", font=("Segoe UI", 14, "bold"), command=self.Them)
            self.btn_Them.place(x=20, y = 315)            
    #Sửa
            self.btn_Sua = CTkButton(self.frameTop, width=70, height=25, text="✍️ Sửa",
                                    fg_color="#6A138D", font=("Segoe UI", 14, "bold"), command=self.Sua)
            self.btn_Sua.place(x=190, y = 315)
    #Xuat Excel
            self.btn_XuatExcel = CTkButton(self.frameTop, width=120, height=25, text="📤 Xuất Excel", command=self.XuatExcel)
            self.btn_XuatExcel.place(x=350, y = 315)
    #Xóa
        self.btn_Xoa = CTkButton(self.frameTop, width=70, height=25, text="🗑️Xóa",
                                    fg_color="#8D1313", font=("Segoe UI", 14, "bold"), command=self.Xoa)
        self.btn_Xoa.place(x=100, y = 315)
    #Lưu 
        self.btn_Luu = CTkButton(self.frameTop, width=70, height=25, text="♻️ Lưu",
                                    fg_color="#132F8D", font=("Segoe UI", 14, "bold"), command=self.Luu)
        self.btn_Luu.place(x=270, y = 315)
    
    def XuatExcel(self):
        # Lấy dữ liệu từ Treeview
        rows = [self.tree.item(item, "values") for item in self.tree.get_children()]
        if not rows:
            messagebox.showwarning("Thông báo", "Không có dữ liệu để xuất Excel!")
            return

        # Hỏi người dùng chọn nơi lưu file
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Lưu file Excel"
        )
        if not file_path:
            return  

        # Tạo workbook mới
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách đặt chỗ"

        # Tiêu đề cột
        columns = ["Mã Đặt Chỗ", "Mã Khách Hàng", "Mã Nhân Viên", "Mã Tour",
                "Người Lớn", "Trẻ Em", "Tổng Tiền", "Ngày Đặt", "Trạng Thái"]
        ws.append(columns)

        # Định dạng tiêu đề
        for col in range(1, len(columns) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        # Thêm dữ liệu
        for row in rows:
            ws.append(row)

        # Lưu file
        try:
            wb.save(file_path)
            messagebox.showinfo("Thành công", f"Xuất Excel thành công:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu file Excel:\n{e}")    
            
    def clear_entries(self):
        self.entry_MaDatCho.delete(0, "end")
        self.cb_MaKhachHang.set(" ")
        self.cb_MaNhanVien.set(" ")
        self.cb_MaTour.set(" ")
        self.entry_SoLuongNguoiLon.delete(0, "end")
        self.entry_SoLuongTreEm.delete(0, "end")
        self.cb_TrangThai.set(" ")
        self.entry_TongTien.configure(state="normal")
        self.entry_TongTien.delete(0, "end")
        self.entry_TongTien.configure(state="disabled")
    
    def load_data(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Load ComboBox khách hàng
        try:
            khachhang = self.db.query("SELECT MaKhachHang, HoTen FROM KHACHHANG")
            kh_list = [f"{kh[0]} - {kh[1]}" for kh in khachhang]
            self.cb_MaKhachHang.configure(values=kh_list)
        except:
            pass
        
        # Load ComboBox nhân viên 
        try:
            nhanvien = self.db.query("SELECT MaNhanVien, HoTen FROM NHANVIEN")
            nv_list = [f"{nv[0]} - {nv[1]}" for nv in nhanvien]
            self.cb_MaNhanVien.configure(values=nv_list)
        except:
            pass
        
        # Load ComboBox tour 
        try:
            tour = self.db.query("SELECT MaTour, TenTour FROM TOUR")
            tour_list = [f"{t[0]} - {t[1]}" for t in tour]
            self.cb_MaTour.configure(values=tour_list)
        except:
            pass
        # Load TrangThai
        list_TrangThai = ["Đã đặt", "Chưa đặt"]
        self.cb_TrangThai.configure(values= [])
        self.cb_TrangThai.configure(values=list_TrangThai)
        
        list_Timkiem = ["Mã đặt chỗ", "Mã khách hàng", "Mã nhân viên", "Mã tuyến"]
        self.cb_TimKiem.configure(values= [])
        self.cb_TimKiem.configure(values=list_Timkiem)
        # Load dữ liệu đặt chỗ
        if BaseForm.UserSession.is_user():
            sql = """SELECT MaDatCho,MaKhachHang, MaNhanVien,MaTour, SoLuongNguoiLon, SoLuongTreEm,TongTien, NgayDat,TrangThaiBooking
                    FROM DATCHO
                    WHERE MaKhachHang = ?
                    ORDER BY NgayDat DESC;
            """
            params = (BaseForm.UserSession.current_user,)
        else:
            sql = "SELECT MaDatCho, MaKhachHang, MaNhanVien, MaTour, SoLuongNguoiLon, SoLuongTreEm, TongTien, NgayDat, TrangThaiBooking FROM DATCHO"
            params = ()
        try:
            rows = self.db.query(sql, params)
            if rows:
                for row in rows:
                    ngay_dat = row[7].strftime("%d/%m/%Y") if hasattr(row[7], "strftime") else str(row[7])
                    tong_tien = row[6]
                    self.tree.insert("", "end", values=(
                        row[0], row[1], row[2], row[3], row[4], row[5], 
                        tong_tien, ngay_dat, row[8]
                    ))
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi truy vấn dữ liệu: {e}")
    
    def on_tree_select(self, event):
        self.clear_entries()
        selected_item = self.tree.selection()
        if selected_item:
            values = self.tree.item(selected_item[0], "values")
            
            self.entry_MaDatCho.insert(0, str(values[0].strip()))
            self.cb_MaKhachHang.set(str(values[1].strip()))
            self.cb_MaNhanVien.set(str(values[2].strip()))
            self.cb_MaTour.set(str(values[3].strip()))
            self.entry_SoLuongNguoiLon.insert(0, str(values[4].strip()))
            self.entry_SoLuongTreEm.insert(0, str(values[5].strip()))
            self.date_NgayDat.set_date(values[7].strip())
            self.cb_TrangThai.set(str(values[8].strip()))
            self.entry_TongTien.configure(state="normal")
            self.entry_TongTien.insert(0, str(values[6].strip()))
            self.entry_TongTien.configure(state="disabled")
    
    def Them(self):
        ma_dat_cho = self.entry_MaDatCho.get().strip()
        ma_kh = self.cb_MaKhachHang.get().strip().split(" - ")[0] if self.cb_MaKhachHang.get() else ""
        ma_nv = self.cb_MaNhanVien.get().strip().split(" - ")[0] if self.cb_MaNhanVien.get() else ""
        ma_tour = self.cb_MaTour.get().strip().split(" - ")[0] if self.cb_MaTour.get() else ""
        so_nguoi_lon = self.entry_SoLuongNguoiLon.get().strip()
        so_tre_em = self.entry_SoLuongTreEm.get().strip()
        trang_thai = self.cb_TrangThai.get().strip()
        ngaydat = self.date_NgayDat.get_date()
        
        if not all([ma_dat_cho, ma_kh, ma_nv, ma_tour, so_nguoi_lon]):
            messagebox.showwarning("Cảnh báo", "Vui lòng điền đầy đủ thông tin bắt buộc.")
            return
        
        # Kiểm tra trùng lặp
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            if values[0].strip() == ma_dat_cho:
                messagebox.showwarning("Cảnh báo", "Mã đặt chỗ đã tồn tại.")
                return
        
        # Tạm thời hiển thị "Đang tính..." trong TreeView
        self.tree.insert("", "end", values=(ma_dat_cho, ma_kh, ma_nv, ma_tour, so_nguoi_lon, so_tre_em, " ", ngaydat.strftime("%d/%m/%Y"), trang_thai))
        
        # Lưu vào list để INSERT sau (TongTien = 0, trigger sẽ tính)
        self.list_them.append((ma_dat_cho, ma_kh, ma_nv, ma_tour, so_nguoi_lon, so_tre_em or 0, ngaydat.strftime("%Y-%m-%d"), trang_thai))
        self.clear_entries()
        messagebox.showinfo("Thành công", "Đã thêm đặt chỗ vào danh sách chờ lưu. Tổng tiền sẽ tự động tính khi lưu.")
    
    def Xoa(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn đặt chỗ để xóa.")
            return
        
        ma_dat_cho = self.tree.item(selected_item[0], "values")[0]
        
        for item in self.list_them:
            if item[0] == ma_dat_cho:
                self.list_them.remove(item)
                break
        else:
            self.list_xoa.append(ma_dat_cho)
        
        self.tree.delete(selected_item[0])
        self.clear_entries()
        messagebox.showinfo("Thành công", "Đã xóa đặt chỗ khỏi danh sách chờ lưu.")
    
    def Sua(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn đặt chỗ để sửa.")
            return
        
        ma_dat_cho = self.entry_MaDatCho.get().strip()
        ma_kh = self.cb_MaKhachHang.get().strip().split(" - ")[0] if self.cb_MaKhachHang.get() else ""
        ma_nv = self.cb_MaNhanVien.get().strip().split(" - ")[0] if self.cb_MaNhanVien.get() else ""
        ma_tour = self.cb_MaTour.get().strip().split(" - ")[0] if self.cb_MaTour.get() else ""
        so_nguoi_lon = self.entry_SoLuongNguoiLon.get().strip()
        so_tre_em = self.entry_SoLuongTreEm.get().strip()
        trang_thai = self.cb_TrangThai.get().strip()
        ngaydat = self.date_NgayDat.get_date()
        
        if not ma_dat_cho:
            messagebox.showwarning("Cảnh báo", "Mã đặt chỗ không được để trống.")
            return
        
        original_ma_dat_cho = self.tree.item(selected_item[0], "values")[0]
        if ma_dat_cho != original_ma_dat_cho:
            messagebox.showwarning("Lỗi", "Không được phép thay đổi mã đặt chỗ!")
            return
        
        # Tạm thời hiển thị "Đang tính..." trong TreeView
        self.tree.item(selected_item[0], values=(ma_dat_cho, ma_kh, ma_nv, ma_tour, so_nguoi_lon, so_tre_em, "Đang tính...", ngaydat.strftime("%d/%m/%Y"), trang_thai))
        
        # Lưu vào list để UPDATE sau (trigger sẽ tính TongTien)
        self.list_sua.append((ma_dat_cho, ma_kh, ma_nv, ma_tour, so_nguoi_lon, so_tre_em or 0, ngaydat.strftime("%Y-%m-%d"), trang_thai))
        self.clear_entries()
        messagebox.showinfo("Thành công", "Đã sửa đặt chỗ trong danh sách chờ lưu. Tổng tiền sẽ tự động tính khi lưu.")
    
    def Luu(self):
        cursor = self.db.conn.cursor()
        try:
            # INSERT - Trigger sẽ tự động tính TongTien
            for item in self.list_them:
                sql = """
                INSERT INTO DATCHO (MaDatCho, MaKhachHang, MaNhanVien, MaTour, SoLuongNguoiLon, 
                                  SoLuongTreEm, TongTien, NgayDat, TrangThaiBooking)
                VALUES (?, ?, ?, ?, ?, ?, 0, ?, ?)
                """
                cursor.execute(sql, (item[0], item[1], item[2], item[3], int(item[4]), int(item[5]), item[6], item[7]))
            
            # DELETE
            for ma_dat_cho in self.list_xoa:
                sql = "DELETE FROM DATCHO WHERE MaDatCho = ?"
                cursor.execute(sql, (ma_dat_cho,))
            
            # UPDATE - Trigger sẽ tự động tính lại TongTien
            for item in self.list_sua:
                sql = """
                UPDATE DATCHO 
                SET MaKhachHang=?, MaNhanVien=?, MaTour=?, SoLuongNguoiLon=?, SoLuongTreEm=?, 
                    NgayDat=?, TrangThaiBooking=?
                WHERE MaDatCho=?
                """
                cursor.execute(sql, (item[1], item[2], item[3], int(item[4]), int(item[5]), item[6], item[7], item[0]))
            
            cursor.commit()
            self.list_them.clear()
            self.list_xoa.clear()
            self.list_sua.clear()
            self.load_data()
            messagebox.showinfo("Thành công", "Đã lưu tất cả các thay đổi vào cơ sở dữ liệu. Tổng tiền đã được tự động tính.")
        except Exception as e:
            cursor.rollback()
            messagebox.showerror("Lỗi", f"Lỗi khi lưu dữ liệu: {e}")
            
    def TimKiem(self):
        # Lấy lựa chọn tìm kiếm từ Combobox
        loai_tim = self.cb_TimKiem.get().strip()
        tu_khoa = self.entry_TimKiem.get().strip().lower()  # chuyển về chữ thường để tìm không phân biệt hoa thường

        if not tu_khoa:
            messagebox.showwarning("Thông báo", "Vui lòng nhập từ khóa để tìm kiếm!")
            return

        # Xóa dữ liệu cũ trên Treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Lấy tất cả dữ liệu từ database
        sql = "SELECT MaDatCho, MaKhachHang, MaNhanVien, MaTour, SoLuongNguoiLon, SoLuongTreEm, TongTien, NgayDat, TrangThaiBooking FROM DATCHO"
        try:
            rows = self.db.query(sql)
            if rows:
                ketqua = []
                for row in rows:
                    # Chọn cột để so sánh dựa trên Combobox
                    if loai_tim == "Mã đặt chỗ":
                        cot_so_sanh = str(row[0])  # DiaDiem
                    elif loai_tim == "Mã khách hàng":
                        cot_so_sanh = str(row[1])  # TenTour
                    elif loai_tim == "Số mã nhân viên":
                        cot_so_sanh = str(row[2])  # ThoiLuong
                    elif loai_tim == "Mã tuyến":
                        cot_so_sanh = str(row[3])
                    else:
                        cot_so_sanh = ""

                    # So sánh từ khóa
                    if tu_khoa in cot_so_sanh.lower():
                        ketqua.append(row)

                # Hiển thị kết quả
                for row in ketqua:
                    ngay_dat = row[7].strftime("%d/%m/%Y") if hasattr(row[7], "strftime") else str(row[7])
                    tong_tien = row[6]
                    self.tree.insert("", "end", values=(
                        row[0], row[1], row[2], row[3], row[4], row[5], 
                        tong_tien, ngay_dat, row[8]
                    ))
            else:
                messagebox.showinfo("Thông báo", "Không có dữ liệu trong cơ sở dữ liệu.")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi truy vấn dữ liệu:\n{e}")